#!/usr/bin/env python3
"""Import the exhibition content: FAQ workbook + exhibitor workbook.

    .venv/bin/python scripts/import-content.py --faq X.xlsx --companies Y.xlsx
    .venv/bin/python scripts/import-content.py ... --apply      # writes

WHAT LANDS WHERE
----------------
FAQ workbook (پرسش/پاسخ): each row → one dataset entry (id faq-<n>) + every
  newline-separated question variant in the cell → one curated question row.
  Multi-question cells are the workbook's own way of saying "these all ask
  the same thing" — one entry, several anchors.
  The sheet arrives in TWO shapes, told apart by the header row (the file is
  the truth, never a flag): the plain پرسش/پاسخ one, and the elecomp one
  with a شماره ویدئو FAQ column first — the number maps the row to its
  FAQ-<n>.mp4 clip (id faq-<number>, stable across re-imports).

Exhibitor workbook (20 columns): each row → one `companies` row (see
  migrations/0013_companies.sql — companies are no longer `dataset` rows):
  * نام/نام(انگلیسی)/درباره/درباره(انگلیسی) (the public answer — the only
                        fields the chatbot may ever serve) plus video_url when
                        the workbook carries booth numbers
  * contact/comms/address/classification fields (the profile columns)
  * curated anchors:    4 Persian templates + 1 English per company, because
                        measured retrieval on company names is exactly where
                        the corpus without anchors failed
  * NO company_leads:   spreadsheet data is not consent; a profile never
                        owns a company (search_companies must keep showing it)

  The organizer sends this workbook in TWO shapes: the plain 20-column one,
  and the same file with a booth-video-number column prepended. There is no
  flag for it — the header row says which shape the file is, so the file
  itself is the source of truth and a wrong flag can never silently shift
  every column by one.

  The elecomp delivery is a THIRD shape, four columns only: booth video
  number, Persian name, the on-video text, the chatbot text. Detected from
  the header too (نام شرکت in column 2 — the 20-column shapes never have
  it there). The chatbot column is the stored text (the on-video column is
  the overlay's wording, kept for the video team); the id is co-<booth
  number>, so a re-import after row reordering still updates in place.
  The description's «در زمینه … فعالیت» phrase is lifted into
  activity_field — the only category signal in this sheet, and the one
  column the company-list tier filters on.

SAFETY
------
Dry-run by default: validates, reports, writes nothing. --apply performs
row-level upserts (existing rows with other ids are untouched), rebuilds the
retrieval index, and prints the follow-ups (eval, golden-set expansion).
"""
import argparse
import os
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# Same pin as run_eval.py, for the same reason: this tool is SQLite-driven
# here and the app imports must not pick a production backend by accident.
os.environ.setdefault("DB_BACKEND", "sqlite")
# An import must never SEED, only import: init_db() with the default flag
# would squeeze the INOTEX starter dataset into whatever database this tool
# is pointed at — on a customer install (elecomp) that is cross-customer
# content pollution, plus a default-credential admin row. Whoever wants the
# starter content can run the app once without this script.
os.environ.setdefault("SEED_DEFAULT_CONTENT", "false")

import openpyxl  # noqa: E402

# The one source of truth for how a stored video_url is spelled. The admin
# upload endpoint builds exactly VIDEO_BASE_URL + "/" + filename
# (app/routers/dataset.py), and /chat hands the stored string straight back to
# the player, so an import that spells it differently would break playback.
from app.config import VIDEO_BASE_URL  # noqa: E402


def _cell(v):
    if v is None:
        return ""
    return re.sub(r"_x000D_|\r", "", str(v)).strip()


def _slug(text: str) -> str:
    text = unicodedata.normalize("NFKD", text or "")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text[:50]


def faq_has_video_column(header) -> bool:
    """True when the FAQ sheet carries the video-number column first.

    Read from the FILE like has_video_column(): the elecomp FAQ sheet ships
    with «شماره ویدئو FAQ» prepended, the older one does not, and guessing
    wrong would read the number as the question.
    """
    first = _cell(header[0] if header else "")
    return "ویدئو" in first or "video" in first.lower()


# The FAQ clips are delivered as FAQ-02.mp4, hyphenated and zero-padded below
# 100 — same tolerance as the booth pattern: the number is the key, not the
# spelling.
FAQ_VIDEO_RE = re.compile(r"faq-?0*(\d+)\.mp4$", re.IGNORECASE)


def load_faq(path: str, videos=None):
    """[(faq_id, title, text, [question_variants], video_url)] + report.

    `videos` is scan_videos(..., FAQ_VIDEO_RE) — {number: filename} — or None
    when the video directory was not available. A row whose number has no
    file gets an empty video_url and a named warning, never a broken player.
    """
    rows, errors, notes, warnings = [], [], [], []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    video_col, with_video, used_numbers = False, 0, set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            video_col = faq_has_video_column(row)
            continue
        if video_col:
            number = _video_number(row[0]) if row else None
            q_raw = _cell(row[1]) if len(row) > 1 else ""
            a = _cell(row[2]) if len(row) > 2 else ""
        else:
            number = None
            q_raw = _cell(row[0]) if row else ""
            a = _cell(row[1]) if len(row) > 1 else ""
        variants = [v.strip() for v in re.split(r"[\n\r]+", q_raw) if v.strip()]
        if not variants and not a:
            continue  # fully empty row
        if not variants:
            errors.append(f"faq row {i}: no question (answer discarded)")
            continue
        if not a:
            errors.append(f"faq row {i}: empty answer — skipped ({variants[0][:40]!r})")
            continue
        if "out of sco" in variants[0].lower():
            notes.append(f"faq row {i}: refusal pair imported as-is ({variants[0][:40]!r})")
        video_url = ""
        if video_col:
            if number is None:
                warnings.append(f"faq row {i} ({variants[0][:40]!r}): no video number")
            elif videos is None:
                warnings.append(f"faq row {i} ({variants[0][:40]!r}): video {number} "
                                f"not verified (no video directory)")
            elif number in videos:
                video_url = f"{VIDEO_BASE_URL}/{videos[number]}"
                used_numbers.add(number)
                with_video += 1
            else:
                warnings.append(f"faq row {i} ({variants[0][:40]!r}): video {number} "
                                f"has no file (expected FAQ-{number:02d}.mp4)")
        fid = f"faq-{number:02d}" if number is not None else f"faq-{i:02d}"
        rows.append((fid, variants[0][:120], a, variants, video_url))
    report = {
        "has_column": video_col,
        "with_video": with_video,
        "warnings": warnings,
        "orphans": (sorted(set(videos) - used_numbers)
                    if video_col and videos is not None else []),
        "files": videos or {},
    }
    return rows, errors, notes, report


FA_CANONICAL = [
    "شرکت {n} چیست؟",
    "درباره {n}",
    "{n} چه کاری انجام می‌دهد؟",
    "فعالیت {n} چیست؟",
]

# Exhibitor workbook column order (verified against the file, 2026-08-26).
COMPANY_COLS = [
    "username", "contact_name", "contact_position", "name", "name_en",
    "email", "website", "phone", "address", "address_en", "fax", "province",
    "company_type", "about", "about_en", "org_stage", "activity_field",
    "participation", "notes", "mobile",
]
PROFILE_MAP = {  # workbook key → companies profile column
    "contact_name": "contact_name", "contact_position": "contact_position",
    "email": "email", "website": "website", "phone": "company_phone",
    "fax": "fax", "address": "address", "address_en": "address_en",
    "province": "province", "company_type": "company_type",
    "org_stage": "org_stage", "activity_field": "activity_field",
    "participation": "participation", "notes": "notes",
}


# The booth videos are delivered as ghorfe-<number>.mp4, zero-padded to two
# digits below 100. One file in the 2026 batch is named ghorfe88.mp4 with no
# hyphen, so the pattern is tolerant of the hyphen and of any padding. The
# number, not the spelling, is the key.
VIDEO_RE = re.compile(r"ghorfe-?0*(\d+)\.mp4$", re.IGNORECASE)


def scan_videos(video_dir: str, pattern=VIDEO_RE):
    """{number: filename} for one directory, or None if it is absent.

    `pattern` picks the delivery family — the booth clips (VIDEO_RE) or the
    FAQ clips (FAQ_VIDEO_RE); the two live in the same media folder.
    None means "existence unknown" — the caller must not invent URLs then.
    A machine without the 6.5 GB of video must still be able to run the
    import.
    """
    if not video_dir or not os.path.isdir(video_dir):
        return None
    found = {}
    for name in sorted(os.listdir(video_dir)):
        m = pattern.search(name)
        if m:
            found[int(m.group(1))] = name
    return found


def _video_number(value) -> int | None:
    """Booth number from a spreadsheet cell. Numeric cells arrive as floats."""
    text = _cell(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def has_video_column(header) -> bool:
    """True when the header row starts with the booth-video-number column.

    Read from the FILE, never from a command-line flag: the organizer sends
    both shapes of this workbook, and guessing wrong would shift all 20
    columns by one and import garbage under the right-looking names.
    """
    return "video" in _cell(header[0] if header else "").lower()


def is_elecomp_companies_sheet(header) -> bool:
    """True for the 4-column elecomp exhibitor sheet.

    Column 2 is exactly «نام شرکت» — the 20-column shapes put the contact's
    full name («نام و نام خانوادگی») there, video column or not, so this one
    header cell tells the two families apart without a flag.
    """
    cells = [_cell(c) for c in (header or [])]
    return len(cells) >= 2 and "نام شرکت" in cells[1]


# The elecomp description is a fixed formula: «... در زمینه <field> فعالیت
# می‌کند ...». The field phrase is the only category data this 4-column
# sheet carries, and the company-list tier (app/services/company_search.py)
# reads exactly activity_field — without it, «شرکت‌های هوش مصنوعی» cannot be
# answered from the companies table at all. Rows the formula does not match
# simply get no field; the facet builder drops overlong phrases on its own.
ELECOMP_FIELD_RE = re.compile(r"در\s+زمینه\s+(.+?)\s+فعالیت")


def _elecomp_activity_field(text: str) -> str:
    m = ELECOMP_FIELD_RE.search(text or "")
    if not m:
        return ""
    return m.group(1).strip(" ،,؛;.")[:70]


def load_companies(path: str, videos=None):
    """[(dataset_id, dataset_fields, profile_fields, anchors)] + report.

    `videos` is scan_videos()'s mapping, or None when the video directory was
    not available. Companies whose file is missing get an empty video_url and
    a named warning — a URL to a file that does not exist would show the
    visitor a broken player.
    """
    rows, errors, warnings = [], [], []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    seen_titles = {}
    video_col, elecomp, with_video = False, False, 0
    seen_numbers, used_numbers = {}, set()

    def attach_booth_video(name, number, i):
        """One video policy for every sheet shape: the number column names
        the clip, a missing file is a warning, a stolen number an error."""
        nonlocal with_video
        url = ""
        if number is None:
            warnings.append(f"{name}: no booth video number in the sheet")
        elif number in seen_numbers:
            errors.append(f"company row {i}: booth video number {number} "
                          f"already taken by row {seen_numbers[number]} — "
                          f"{name!r} gets no video")
        elif videos is None:
            warnings.append(f"{name}: booth video {number} not verified "
                            f"(no video directory)")
        elif number in videos:
            # Exactly the shape app/routers/dataset.py writes on upload:
            # VIDEO_BASE_URL + "/" + the file's own name.
            url = f"{VIDEO_BASE_URL}/{videos[number]}"
            used_numbers.add(number)
            with_video += 1
        else:
            warnings.append(f"{name}: booth video {number} has no file "
                            f"(expected ghorfe-{number:02d}.mp4)")
        if number is not None:
            seen_numbers.setdefault(number, i)
        return url

    def dedupe_name(name, i):
        key = re.sub(r"\s+", " ", name)
        if key in seen_titles:
            errors.append(f"company row {i}: duplicate name {key!r} "
                          f"(first at row {seen_titles[key]})")
            return None
        seen_titles[key] = i
        return key

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            video_col = has_video_column(row)
            elecomp = is_elecomp_companies_sheet(row)
            continue

        if elecomp:
            # 4-column elecomp sheet: number | نام شرکت | ویدیو چت‌بات | چت بات.
            # The chatbot column is the answer a visitor is served; the
            # on-video column is the overlay's wording and only a fallback.
            number = _video_number(row[0]) if row else None
            name = _cell(row[1]) if len(row) > 1 else ""
            video_text = _cell(row[2]) if len(row) > 2 else ""
            chat_text = _cell(row[3]) if len(row) > 3 else ""
            if not name:
                continue
            if dedupe_name(name, i) is None:
                continue
            video_url = attach_booth_video(name, number, i)
            cid = f"co-{number:03d}" if number is not None else f"co-{i:03d}"
            anchors = [t.format(n=name) for t in FA_CANONICAL]
            profile = {}
            if number is not None:
                profile["booth_number"] = str(number)
            field = _elecomp_activity_field(chat_text or video_text)
            if field:
                profile["activity_field"] = field
            rows.append((cid, {"id": cid, "title": name, "title_en": "",
                               "text": chat_text or video_text, "text_en": "",
                               "video_url": video_url},
                         profile, anchors))
            continue

        if video_col:
            number = _video_number(row[0])
            cells = row[1:]
        else:
            number, cells = None, row
        c = dict(zip(COMPANY_COLS, (_cell(v) for v in cells)))
        if not c["name"] and not c["name_en"] and not c["about"]:
            continue
        if not c["name"]:
            errors.append(f"company row {i}: no Persian name — skipped")
            continue
        if dedupe_name(c["name"], i) is None:
            continue

        video_url = ""
        if video_col:
            video_url = attach_booth_video(c["name"], number, i)

        cid = _slug(c["name_en"]) or f"co-{i:03d}"
        mobile = c["mobile"] or c["username"]
        mobile = re.sub(r"^(98|0098)", "0", mobile) if mobile else ""
        profile = {dst: c[src] for src, dst in PROFILE_MAP.items() if c.get(src)}
        if mobile:
            profile["contact_mobile"] = mobile
        # Same number the video lookup above already parsed from column 0 —
        # free, since this sheet shape names booth number and video by
        # definition (has_video_column()). The plain-workbook shape (no video
        # column) has no number to give here; booth_number stays unset and an
        # admin fills it in later.
        if number is not None:
            profile["booth_number"] = str(number)
        anchors = [t.format(n=c["name"]) for t in FA_CANONICAL]
        if c["name_en"]:
            anchors.append(f"What is {c['name_en']}?")
        rows.append((cid, {"id": cid, "title": c["name"], "title_en": c["name_en"],
                           "text": c["about"], "text_en": c["about_en"],
                           "video_url": video_url},
                     profile, anchors))

    report = {
        "has_column": video_col or elecomp,
        "with_video": with_video,
        "warnings": warnings,
        "orphans": (sorted(set(videos) - used_numbers)
                    if (video_col or elecomp) and videos is not None else []),
        "files": videos or {},
    }
    return rows, errors, report


def main() -> int:
    p = argparse.ArgumentParser(description="Import FAQ + exhibitor workbooks.")
    p.add_argument("--faq", help="FAQ workbook (پرسش/پاسخ, with or without "
                                 "the video-number column)")
    p.add_argument("--companies",
                    help="exhibitor workbook (20 columns, 21 with the "
                         "booth-video-number column first, or the 4-column "
                         "elecomp sheet)")
    p.add_argument("--video-dir", default="media/videos",
                    help="where the booth + FAQ videos live — read only, to "
                         "check which files exist (default: media/videos)")
    p.add_argument("--apply", action="store_true",
                    help="actually write; without it everything is a dry-run")
    args = p.parse_args()
    if not args.faq and not args.companies:
        sys.exit("nothing to do: pass --faq and/or --companies")

    videos = scan_videos(args.video_dir)
    faq_videos = scan_videos(args.video_dir, FAQ_VIDEO_RE)
    if videos is None and faq_videos is None:
        print(f"NOTE: video directory {args.video_dir!r} not found — "
              f"skipping the file-existence check. No video will be attached.")
    else:
        print(f"Booth video files found:  {len(videos or {})} in {args.video_dir}")
        print(f"FAQ video files found:    {len(faq_videos or {})} in {args.video_dir}")

    empty_faq_report = {"has_column": False, "with_video": 0, "warnings": [],
                        "orphans": [], "files": faq_videos or {}}
    faq, faq_errors, faq_notes, faq_video_report = (
        load_faq(args.faq, faq_videos) if args.faq else ([], [], [], empty_faq_report))
    companies, co_errors, video_report = (
        load_companies(args.companies, videos) if args.companies
        else ([], [], {"has_column": False, "with_video": 0, "warnings": [],
                       "orphans": [], "files": videos or {}}))

    os.environ.setdefault("OPENAI_API_KEY", "import")
    from app.db import connection as dbconn
    from app.db.connection import get_db_connection
    # An import must not create login accounts. init_db() unconditionally
    # seeds an admin (correct for a fresh install's first boot, wrong for a
    # tool aimed at a live customer database — it left a default-username
    # row with an auto-generated password behind, i.e. an account nobody
    # can audit). Disabled here, restored right after.
    _real_seed_admin = dbconn._seed_admin
    dbconn._seed_admin = lambda cursor: None
    # Same rule for the starter CONTENT, set on the module attribute and
    # not just the env: init_db() reads config.SEED_DEFAULT_CONTENT at call
    # time, and anything that imported app.config before this script (a
    # test, a wrapper) has already latched the env default to True.
    import app.config as appconfig
    _real_seed_content = appconfig.SEED_DEFAULT_CONTENT
    appconfig.SEED_DEFAULT_CONTENT = False
    try:
        dbconn.init_db()
    finally:
        dbconn._seed_admin = _real_seed_admin
        appconfig.SEED_DEFAULT_CONTENT = _real_seed_content
    conn = get_db_connection()
    existing = {r["id"] for r in conn.execute("SELECT id FROM dataset").fetchall()}
    existing_companies = {r["id"] for r in conn.execute("SELECT id FROM companies").fetchall()}
    existing_q = {r["question"] for r in conn.execute("SELECT question FROM questions").fetchall()}
    conn.close()

    faq_clash = [r[0] for r in faq if r[0] in existing]
    co_clash = [r[0] for r in companies if r[0] in existing_companies]
    # `dataset` and `companies` are two separate primary-key spaces
    # (migrations/0013_companies.sql) — nothing stops an id from existing in
    # both. That is never an upsert, it is a real conflict: whichever id wins
    # `dataset_lookup.get(id) or companies_lookup.get(id)` would silently
    # shadow the other row. Caught here, before either INSERT runs.
    cross_clash = ([r[0] for r in faq if r[0] in existing_companies]
                   + [c[0] for c in companies if c[0] in existing])

    print(f"FAQ rows to import:      {len(faq)}")
    for fid, title, _, variants, _vurl in faq[:5]:
        print(f"   {fid}: {title[:60]!r} ({len(variants)} question variants)")
    if len(faq) > 5:
        print(f"   … and {len(faq) - 5} more")
    for e in faq_errors:
        print(f"   ERROR {e}")
    for n in faq_notes:
        print(f"   NOTE  {n}")
    if faq_video_report["has_column"]:
        print(f"FAQ entries with a video: {faq_video_report['with_video']}")
        for w in faq_video_report["warnings"]:
            print(f"   WARNING {w}")
        for n in faq_video_report["orphans"]:
            print(f"   WARNING FAQ video {n} "
                  f"({faq_video_report['files'][n]}) matches no FAQ row")
    print(f"Companies to import:     {len(companies)}")
    for cid, ds, profile, anchors in companies[:5]:
        print(f"   {cid}: {ds['title'][:40]!r} profile_fields="
              f"{sum(1 for v in profile.values() if v)} anchors={len(anchors)}")
    if len(companies) > 5:
        print(f"   … and {len(companies) - 5} more")
    for e in co_errors:
        print(f"   ERROR {e}")

    if video_report["has_column"]:
        without = video_report["warnings"]
        print(f"Companies with a video:  {video_report['with_video']}")
        print(f"Companies without one:   {len(without)}")
        for w in without:
            print(f"   WARNING {w}")
        print(f"Videos with no company:  {len(video_report['orphans'])}")
        for n in video_report["orphans"]:
            print(f"   WARNING booth video {n} "
                  f"({video_report['files'][n]}) matches no company row")
    elif args.companies:
        print("Booth video column:      absent — "
              "existing video_url values are left untouched")

    print(f"Id collisions with DB:   faq={len(faq_clash)} companies={len(co_clash)}")
    if faq_clash or co_clash:
        print(f"   {faq_clash + co_clash}")
    if cross_clash:
        print(f"   ERROR: id(s) collide ACROSS dataset and companies "
              f"(two different tables, same id): {cross_clash}")

    if cross_clash:
        sys.exit("Aborting: fix the colliding id(s) in the workbook first — "
                 "an id must not name both a dataset row and a company.")

    if not args.apply:
        print("\nDRY RUN — nothing written. Re-run with --apply to import.")
        return 0

    from app.db.connection import get_db_connection
    from app.services import company_profiles
    conn = get_db_connection()
    n_faq = n_co = n_q = 0

    # Pass 1 — every FAQ (dataset) and company row first, committed.
    # upsert_profile opens its OWN connection, and a second connection cannot
    # see rows this one has not committed yet (the first --apply died exactly
    # there: profile upsert said «این شرکت در دانش‌نامه نیست» for a row
    # sitting uncommitted one connection over).
    for fid, title, text, variants, video_url in faq:
        # Same no-wipe rule as companies: a re-import run without the video
        # directory must not blank a video that is already attached.
        conn.execute(
            "INSERT INTO dataset (id, title, text, video_url, title_en, text_en)"
            " VALUES (?, ?, ?, ?, '', '')"
            " ON CONFLICT(id) DO UPDATE SET title=excluded.title, text=excluded.text,"
            " video_url=CASE WHEN excluded.video_url != ''"
            " THEN excluded.video_url ELSE dataset.video_url END",
            (fid, title, text, video_url))
        n_faq += 1
    for cid, ds, profile, anchors in companies:
        # A company is one `companies` row, not a `dataset` row —
        # migrations/0013_companies.sql. video_url is updated ONLY when this
        # run actually has one. A re-import from the plain 20-column workbook
        # (or from a machine without the video files) must never wipe a video
        # that is already attached.
        conn.execute(
            "INSERT INTO companies (id, title, text, video_url, title_en, text_en)"
            " VALUES (?, ?, ?, ?, ?, ?)"
            " ON CONFLICT(id) DO UPDATE SET title=excluded.title, text=excluded.text,"
            " title_en=excluded.title_en, text_en=excluded.text_en,"
            " video_url=CASE WHEN excluded.video_url != ''"
            " THEN excluded.video_url ELSE companies.video_url END",
            (cid, ds["title"], ds["text"], ds["video_url"],
             ds["title_en"], ds["text_en"]))
        n_co += 1
    conn.commit()

    # Pass 2 — profiles, now over rows that are visible to every connection.
    # Before ANY question insert: upsert_profile opens its own connection, and
    # an open write transaction on this one locks it out (SQLite allows a
    # single writer). Interleaved, the second company waited out the 5s
    # busy_timeout and died with "database is locked".
    for cid, ds, profile, anchors in companies:
        if profile:
            company_profiles.upsert_profile(cid, profile, source="import")

    # Pass 3 — curated anchors and FAQ question variants, one writer, one
    # transaction.
    for fid, title, text, variants, _vurl in faq:
        for v in variants:
            if v in existing_q:
                continue
            conn.execute("INSERT INTO questions (question, dataset_id, video_url)"
                         " VALUES (?, ?, '')", (v, fid))
            n_q += 1
    for cid, ds, profile, anchors in companies:
        for a in anchors:
            if a in existing_q:
                continue
            conn.execute("INSERT INTO questions (question, dataset_id, video_url)"
                         " VALUES (?, ?, '')", (a, cid))
            n_q += 1
    conn.commit()
    conn.close()

    from app.services.search import reindex_and_publish
    reindex_and_publish()
    print(f"\nAPPLIED: {n_faq} FAQ rows, {n_co} companies, "
          f"{n_q} curated questions, "
          f"{video_report['with_video']} booth videos, "
          f"{faq_video_report['with_video']} FAQ videos. Index rebuilt.")
    print("Next: .venv/bin/python scripts/run_eval.py   # re-baseline")
    return 0


if __name__ == "__main__":
    sys.exit(main())
