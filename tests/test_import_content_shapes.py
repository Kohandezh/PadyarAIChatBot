"""The elecomp sheet shapes: a FAQ sheet with a video-number column and a
4-column exhibitor sheet.

The elecomp delivery differs from every workbook the importer knew
(test_booth_videos.py covers those): the FAQ sheet carries a
«شماره ویدئو FAQ» column first — the number names the FAQ-<n>.mp4 clip — and
the exhibitor sheet is four columns only (booth number, Persian name, the
on-video text, the chatbot text). Both are told apart from the OTHER shapes
by the header row, because a flag can be passed wrong and the file cannot.

WHAT IS UNDER TEST

  * the FAQ number column maps the row to its FAQ-<n>.mp4 and to a stable
    faq-<number> id (a re-import after row reordering updates in place)
  * the chatbot column — not the on-video column — is the stored company
    text, because the chatbot serves it and the overlay wording is the
    video team's
  * an empty answer is still an import error, not a silent empty entry
  * the write path stores the FAQ video_url, with the same don't-wipe rule
    the booth videos already had
"""
import importlib.util
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
IMPORTER = ROOT / "scripts" / "import-content.py"

FAQ_HEADER = ["شماره ویدئو FAQ", "پرسش", "پاسخ"]
ELECOMP_HEADER = ["شماره ویدئو شرکت", "نام شرکت", "ویدیو چت‌بات", "چت بات"]


@pytest.fixture
def importer():
    """scripts/import-content.py loaded by path — its name has a hyphen, so
    it cannot be imported the normal way."""
    spec = importlib.util.spec_from_file_location("_import_content", IMPORTER)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _faq_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(FAQ_HEADER)
    ws.append([2.0, "الکامپ چیست؟", "نمایشگاه الکترونیک است."])
    ws.append([3.0, "سؤال یک؟\nسؤال دو؟", "پاسخ با دو پرسش."])
    ws.append([49.0, "پرسش بدون پاسخ؟", ""])
    wb.save(path)
    return str(path)


def _elecomp_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ELECOMP_HEADER)
    ws.append([2.0, "فرزام", "متن روی ویدیو فرزام",
               "فرزام در زمینه امنیت شبکه فعالیت می‌کند."])
    ws.append([3.0, "ستاره", "متن روی ویدیو ستاره",
               "ستاره در زمینه هوش مصنوعی فعالیت می‌کند."])
    # Booth 9 has no file in the fixture directory on purpose.
    ws.append([9.0, "بی‌فایل", "متن روی ویدیو", "بی‌فایل توضیح دیگری دارد."])
    wb.save(path)
    return str(path)


def _video_dir(tmp_path):
    """The stand-in media folder: booth clips and FAQ clips together."""
    d = tmp_path / "videos"
    d.mkdir(exist_ok=True)
    for name in ("ghorfe-02.mp4", "ghorfe-03.mp4", "ghorfe-12.mp4",
                 "FAQ-02.mp4", "FAQ-03.mp4", "FAQ-49.mp4"):
        (d / name).write_bytes(b"fake mp4 payload")
    return str(d)


# ── FAQ sheet ────────────────────────────────────────────────────────────

def test_the_faq_number_column_names_the_id_and_the_clip(importer, tmp_path):
    xlsx = _faq_book(tmp_path / "faq.xlsx")
    videos = importer.scan_videos(_video_dir(tmp_path), importer.FAQ_VIDEO_RE)
    rows, errors, _notes, report = importer.load_faq(xlsx, videos)

    assert any("empty answer" in e for e in errors)  # row 49, no answer
    by_id = {fid: (text, vurl) for fid, _t, text, _v, vurl in rows}
    # The stable id is the NUMBER, so re-imports after reordering update
    # in place instead of duplicating the entry.
    assert "/media/videos/FAQ-02.mp4" == by_id["faq-02"][1]
    assert "/media/videos/FAQ-03.mp4" == by_id["faq-03"][1]
    assert report["with_video"] == 2


def test_a_multi_question_cell_is_still_one_entry_several_anchors(
        importer, tmp_path):
    xlsx = _faq_book(tmp_path / "faq.xlsx")
    rows, _errors, _notes, _report = importer.load_faq(xlsx, None)

    entry = [r for r in rows if r[0] == "faq-03"][0]
    assert entry[3] == ["سؤال یک؟", "سؤال دو؟"]


def test_an_empty_answer_is_an_error_not_an_empty_entry(importer, tmp_path):
    xlsx = _faq_book(tmp_path / "faq.xlsx")
    videos = importer.scan_videos(_video_dir(tmp_path), importer.FAQ_VIDEO_RE)
    rows, errors, _notes, report = importer.load_faq(xlsx, videos)

    assert [r[0] for r in rows].count("faq-49") == 0
    assert any("empty answer" in e for e in errors)
    # Its clip exists but the row was skipped — named as an orphan, so the
    # operator can tell a skipped row from a missing file.
    assert report["orphans"] == [49]


def test_the_plain_faq_sheet_without_the_video_column_imports_as_before(
        importer, tmp_path):
    """Regression: the older two-column FAQ sheet must keep working."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["پرسش", "پاسخ"])
    ws.append(["پرسش قدیمی؟", "پاسخ قدیمی."])
    wb.save(tmp_path / "plain.xlsx")

    rows, errors, _notes, report = importer.load_faq(str(tmp_path / "plain.xlsx"))
    assert errors == []
    assert report["has_column"] is False
    assert rows[0][0] == "faq-01"
    assert rows[0][4] == ""


# ── elecomp exhibitor sheet ──────────────────────────────────────────────

def test_the_four_column_sheet_parses_number_name_and_chatbot_text(
        importer, tmp_path):
    xlsx = _elecomp_book(tmp_path / "co.xlsx")
    videos = importer.scan_videos(_video_dir(tmp_path))
    rows, errors, report = importer.load_companies(xlsx, videos)

    assert errors == []
    assert report["has_column"] is True
    by_id = {cid: ds for cid, ds, _p, _a in rows}
    # The booth NUMBER is the id: co-002, stable across re-imports.
    assert by_id["co-002"]["title"] == "فرزام"
    # The chatbot column is the stored text, not the on-video column.
    assert by_id["co-002"]["text"] == "فرزام در زمینه امنیت شبکه فعالیت می‌کند."
    assert by_id["co-002"]["video_url"] == "/media/videos/ghorfe-02.mp4"
    assert by_id["co-003"]["video_url"] == "/media/videos/ghorfe-03.mp4"


def test_the_four_column_sheet_still_anchors_and_reports_missing_files(
        importer, tmp_path):
    xlsx = _elecomp_book(tmp_path / "co.xlsx")
    videos = importer.scan_videos(_video_dir(tmp_path))
    rows, _errors, report = importer.load_companies(xlsx, videos)

    by_id = {cid: (ds, profile, anchors) for cid, ds, profile, anchors in rows}
    ds, profile, anchors = by_id["co-009"]
    assert ds["video_url"] == ""  # no file → no broken player
    assert any("بی‌فایل" in w for w in report["warnings"])
    assert profile["booth_number"] == "9"
    assert "شرکت فرزام چیست؟" in by_id["co-002"][2]
    # The clip nobody claimed is named too.
    assert report["orphans"] == [12]


def test_the_twenty_column_shapes_are_still_recognized(importer):
    """The discriminator must not misfire on the older headers."""
    plain = ["نام کاربری", "نام و نام خانوادگی ", "سمت", "نام"]
    with_video = ["Video number nme"] + plain
    assert importer.is_elecomp_companies_sheet(plain) is False
    assert importer.is_elecomp_companies_sheet(with_video) is False
    assert importer.is_elecomp_companies_sheet(ELECOMP_HEADER) is True


def test_the_description_formula_becomes_activity_field(importer, tmp_path):
    """«شرکت‌های هوش مصنوعی» is answered by the company-list tier, which
    filters on activity_field — a column this sheet does not have. The only
    category signal is the «در زمینه … فعالیت» phrase inside the chatbot
    text, so the importer lifts it out. Without this, the list tier sees
    zero facets and no list question can ever be answered."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ELECOMP_HEADER)
    ws.append([5.0, "هوشمند",
               "متن روی ویدیو",
               "هوشمند در زمینه هوش مصنوعی و امنیت شبکه فعالیت می‌کند."])
    ws.append([6.0, "بازرگان", "متن روی ویدیو", "بازرگان کالا وارد می‌کند."])
    wb.save(tmp_path / "f.xlsx")

    rows, _errors, _report = importer.load_companies(str(tmp_path / "f.xlsx"), None)
    profiles = {cid: p for cid, _ds, p, _a in rows}
    assert profiles["co-005"]["activity_field"] == "هوش مصنوعی و امنیت شبکه"
    assert "activity_field" not in profiles["co-006"]  # formula absent → no field


# ── Write path ───────────────────────────────────────────────────────────

def test_applying_stores_the_faq_and_booth_video_urls(importer, monkeypatch,
                                                      tmp_path):
    """The wiring, not the parsing: what the visitor's player receives."""
    import app.config as config
    monkeypatch.setattr(config, "DB_PATH", str(tmp_path / "shapes.db"))
    monkeypatch.setattr(config, "SEED_DEFAULT_CONTENT", False)

    argv = ["import-content.py",
            "--faq", _faq_book(tmp_path / "faq.xlsx"),
            "--companies", _elecomp_book(tmp_path / "co.xlsx"),
            "--video-dir", _video_dir(tmp_path),
            "--apply"]
    monkeypatch.setattr(importer.sys, "argv", argv)
    assert importer.main() == 0

    import app.db.connection as dbc
    conn = dbc.get_db_connection()
    faq = dict(conn.execute(
        "SELECT id, video_url FROM dataset").fetchall())
    co = dict(conn.execute(
        "SELECT id, video_url FROM companies").fetchall())
    questions = {r["question"] for r in
                 conn.execute("SELECT question FROM questions").fetchall()}
    conn.close()

    assert faq["faq-02"] == "/media/videos/FAQ-02.mp4"
    assert faq["faq-03"] == "/media/videos/FAQ-03.mp4"
    assert co["co-002"] == "/media/videos/ghorfe-02.mp4"
    assert co["co-009"] == ""
    assert "الکامپ چیست؟" in questions
    assert "سؤال دو؟" in questions          # the second variant anchored too
    assert "شرکت ستاره چیست؟" in questions  # generated company anchor
